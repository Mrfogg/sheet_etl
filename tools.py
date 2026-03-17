import os
from utils.logger import setup_logger
from openpyxl import load_workbook
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from utils.excel_toolkit import ExcelToolkit, calculate_token_cost_line

logger = setup_logger(__name__)
    
def _generate_sheets_markdown_summary(excel_path: str, sheet_index = 0,total_token_budget: int = 50000) -> []:
    """Generate a markdown summary of all sheets in the workbook."""
    try:
        workbook = load_workbook(excel_path, data_only=True)
        overview_parts = []

        overview_parts.append(f"📊 **Excel File Overview: {os.path.basename(excel_path)}**\n")
        overview_parts.append(f"**Total Sheets:** {len(workbook.sheetnames)}\n")

        # Token budget management
        available_tokens = total_token_budget
        result = []

        # Distribute tokens among sheets
        tokens_per_sheet = available_tokens // len(workbook.sheetnames) if workbook.sheetnames else 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_parts = []

            sheet_parts.append(f"\n**📄 Sheet: '{sheet_name}'**")
            sheet_parts.append(f"- Dimensions: {sheet.max_row} rows × {sheet.max_column} columns")

            if tokens_per_sheet > 0:
                # Get comprehensive preview with token limit
                preview_result = _get_sheet_preview_with_token_limit(
                    sheet,
                    tokens_per_sheet,
                    max_rows=min(sheet.max_row, 10000),  # Cap at 10000 rows for performance
                    max_cols=min(sheet.max_column, 1000)   # Cap at 1000 columns
                )

                sheet_parts.append(f"- Data Preview ({preview_result['rows_shown']} of {sheet.max_row} rows, "
                                f"{preview_result['cols_shown']} of {sheet.max_column} columns):")

                if preview_result['is_truncated']:
                    sheet_parts.append("  ⚠️ Preview truncated to fit token budget")

                # Add data preview in markdown table format with A1 notation
                sheet_parts.append("  Data:")
                markdown_rows = []
                for row_data in preview_result['formatted_data']:
                    markdown_rows.append(f"| {' | '.join(row_data)} |")

                # Join all rows with newline and backslash-n for compact representation
                if markdown_rows:
                    sheet_parts.append("  " + "\\n".join(markdown_rows))

                # Add data summary if we couldn't show all rows
                if preview_result['rows_shown'] < sheet.max_row:
                    sheet_parts.append(f"\n  📊 Sheet Summary:")
                    sheet_parts.append(f"  - Total rows: {sheet.max_row}")
                    sheet_parts.append(f"  - Total columns: {sheet.max_column}")
                    sheet_parts.append(f"  - Rows shown in preview: {preview_result['rows_shown']}")

            result.append("\n".join(overview_parts + sheet_parts))
        return result

    except Exception as e:
        logger.error(f"Error generating Excel overview: {str(e)}")
        return f"❌ Error generating Excel overview: {str(e)}"

def _get_sheet_preview_with_token_limit(sheet, token_budget: int,
                                    max_rows: int = 10000, max_cols: int = 1000) -> Dict[str, Any]:
    """Get a preview of sheet data that fits within a token budget."""
    preview_data = []
    formatted_data = []
    tokens_used = 0
    rows_shown = 0

    start_row = 1

    # Calculate effective limits
    max_data_rows = min(max_rows, sheet.max_row)
    max_data_cols = min(max_cols, sheet.max_column)

    # Iterate through rows and accumulate data within token budget
    for row_idx in range(start_row, max_data_rows + 1):
        row_cells = []
        formatted_row_cells = []

        # Get actual row data
        for col_idx in range(1, max_data_cols + 1):
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
            cell = sheet[cell_ref]
            cell_value = cell.value

            # Format cell value for display
            display_value = str(cell_value) if cell_value is not None else ""

            # Escape markdown special characters
            display_value = display_value.replace("|", "\\|").replace("\n", " ").replace("\r", " ")

            # Simple format: A1:value
            formatted_cell = f"{cell_ref}:{display_value}"

            row_cells.append(cell_value)
            formatted_row_cells.append(formatted_cell)

        # Convert formatted row to string to estimate tokens
        row_str = " | ".join(formatted_row_cells)
        row_tokens = calculate_token_cost_line(row_str)

        # Check if adding this row would exceed budget
        if tokens_used + row_tokens > token_budget:
            # Try to add at least some rows even if over budget for minimal data
            if rows_shown < 5:  # Ensure we show at least 5 rows if possible
                preview_data.append(row_cells)
                formatted_data.append(formatted_row_cells)
                rows_shown += 1
                tokens_used += row_tokens
            break

        preview_data.append(row_cells)
        formatted_data.append(formatted_row_cells)
        rows_shown += 1
        tokens_used += row_tokens

    return {
        'data': preview_data,
        'formatted_data': formatted_data,
        'rows_shown': rows_shown,
        'cols_shown': max_data_cols,
        'start_row': start_row,
        'is_truncated': rows_shown < max_data_rows,
        'tokens_used': tokens_used
    }